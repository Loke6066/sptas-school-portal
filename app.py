import os
import json
import random
import string
import io
from flask import Flask, render_template, jsonify, request, send_file
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'sptas-school-secret-2026')

DB_DIR = os.path.dirname(__file__)

# In-memory OTP store {phone: otp_code}
OTP_STORE = {}

# =========================================================
# DATABASE HELPERS
# =========================================================
def load_json(filename, default=None):
    path = os.path.join(DB_DIR, filename)
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return default if default is not None else []

def save_json(filename, data):
    path = os.path.join(DB_DIR, filename)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except:
        return False

def load_db():       return load_json('students_db.json', [])
def save_db(d):      return save_json('students_db.json', d)
def load_teachers(): return load_json('teachers_db.json', [])
def save_teachers(d): return save_json('teachers_db.json', d)
def load_timetables(): return load_json('timetables_db.json', {})
def save_timetables(d): return save_json('timetables_db.json', d)
def load_settings(): return load_json('settings.json', {})
def save_settings(d): return save_json('settings.json', d)
def load_meetings(): return load_json('meetings_db.json', [])
def save_meetings(d): return save_json('meetings_db.json', d)

def load_activities(): return load_json('activities_db.json', [])
def save_activities(d): return save_json('activities_db.json', d)

def log_activity(role, name, description):
    import time
    activities = load_activities()
    new_act = {
        "id": f"act_{int(time.time())}_{random.randint(1000, 9999)}",
        "role": role,
        "name": name,
        "description": description,
        "timestamp": datetime.now().strftime("%d %b %Y, %I:%M %p")
    }
    activities.insert(0, new_act)
    save_activities(activities[:100])


DEFAULT_GOVT_HOLIDAYS = [
    "2026-01-01", "2026-01-14", "2026-01-15", "2026-01-26",
    "2026-03-18", "2026-03-27", "2026-04-05", "2026-04-14",
    "2026-05-01", "2026-08-15", "2026-09-07", "2026-10-02",
    "2026-10-20", "2026-11-08", "2026-12-25"
]

def is_school_holiday(date_str):
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        return False, ""
    
    settings = load_settings()
    custom_holidays = settings.get("custom_holidays", [])
    custom_working_days = settings.get("custom_working_days", [])
    holiday_reasons = settings.get("holiday_reasons", {})
    
    # Custom working day overrides Sunday/Govt holiday
    if date_str in custom_working_days:
        return False, "Working Day"
        
    # Custom holiday overrides Sunday/Govt holiday
    if date_str in custom_holidays:
        reason = holiday_reasons.get(date_str, "School Holiday")
        return True, reason
        
    # Govt holiday
    if date_str in DEFAULT_GOVT_HOLIDAYS:
        return True, "Govt Holiday"
        
    # Sunday
    if dt.weekday() == 6: # Sunday
        return True, "Sunday"
        
    return False, ""

def clean_phone(phone):
    if not phone: return ""
    return str(phone).replace(" ","").replace("-","").replace("+91","").strip()

# =========================================================
# ROUTES
# =========================================================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/stats', methods=['GET'])
def get_stats():
    students = load_db()
    teachers = load_teachers()
    
    teacher_id = request.args.get('teacher_id', '').strip()
    if teacher_id:
        target_teacher = None
        for t in teachers:
            if t["id"] == teacher_id:
                target_teacher = t
                break
        if target_teacher:
            profile = dict(target_teacher)
            t_classes = profile.get("classes")
            if not t_classes and profile.get("class") and profile.get("section"):
                t_classes = [f"{profile['class']}-{profile['section']}"]
            elif not t_classes:
                t_classes = []
            
            filtered = []
            for s in students:
                s_key = f"{s.get('class','')}-{s.get('section','')}"
                if s_key in t_classes:
                    filtered.append(s)
            students = filtered

    total = len(students)
    classes = len(set(f"{s['class']}-{s['section']}" for s in students)) if students else 0
    avg_att = f"{sum(s['current_status']['attendance_percentage'] for s in students)/total:.1f}%" if total else "0%"
    
    # Today's attendance ratio
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_present = 0
    for s in students:
        rec = next((r for r in s.get("attendance_records", []) if r["date"] == today_str), None)
        if rec:
            if rec["status"] == "Present":
                today_present += 1
            elif rec["status"] == "Half Day":
                today_present += 0.5

    return jsonify({
        "total_students": total,
        "active_classes": classes,
        "overall_attendance": avg_att,
        "teachers": len(teachers),
        "today_present": today_present,
        "today_total": total
    })

# =========================================================
# SETTINGS
# =========================================================
@app.route('/api/settings', methods=['GET'])
def get_settings():
    s = load_settings()
    return jsonify({"success": True, "principal_name": s.get("principal_name","Principal"),
                    "school_name": s.get("school_name","Silverwood International School")})

@app.route('/api/settings/update', methods=['POST'])
def update_settings():
    data = request.get_json() or {}
    s = load_settings()
    if "principal_name" in data: s["principal_name"] = data["principal_name"].strip()
    if "school_name" in data: s["school_name"] = data["school_name"].strip()
    if save_settings(s):
        return jsonify({"success": True})
    return jsonify({"success": False}), 500

def get_sundays_of_current_month():
    now = datetime.now()
    year = now.year
    month = now.month
    import calendar
    c = calendar.Calendar(firstweekday=calendar.SUNDAY)
    sundays = []
    for day in c.itermonthdates(year, month):
        if day.month == month and day.weekday() == 6: # Sunday
            sundays.append(day.strftime("%Y-%m-%d"))
    return sorted(sundays)

@app.route('/api/holidays', methods=['GET'])
def get_holidays():
    settings = load_settings()
    return jsonify({
        "default_govt_holidays": DEFAULT_GOVT_HOLIDAYS,
        "custom_holidays": settings.get("custom_holidays", []),
        "custom_working_days": settings.get("custom_working_days", []),
        "holiday_reasons": settings.get("holiday_reasons", {}),
        "holiday_messages": settings.get("holiday_messages", {}),
        "sundays": get_sundays_of_current_month()
    })

@app.route('/api/holidays/check', methods=['GET'])
def check_holiday():
    date_str = request.args.get('date', '').strip()
    if not date_str:
        date_str = datetime.now().strftime("%Y-%m-%d")
    is_h, reason = is_school_holiday(date_str)
    settings = load_settings()
    holiday_messages = settings.get("holiday_messages", {})
    msg = holiday_messages.get(date_str, "")
    return jsonify({"is_holiday": is_h, "reason": reason, "message": msg})

@app.route('/api/holidays/set', methods=['POST'])
def set_holiday_status():
    data = request.get_json() or {}
    date_str = data.get("date", "").strip()
    status = data.get("status", "").strip() # "holiday", "working", or "clear"
    reason = data.get("reason", "").strip() or "School Holiday"
    message = data.get("message", "").strip()
    if not date_str:
        return jsonify({"success": False, "message": "Date is required."}), 400
        
    settings = load_settings()
    custom_holidays = set(settings.get("custom_holidays", []))
    custom_working_days = set(settings.get("custom_working_days", []))
    holiday_reasons = settings.get("holiday_reasons", {})
    holiday_messages = settings.get("holiday_messages", {})
    
    # Reset existing
    custom_holidays.discard(date_str)
    custom_working_days.discard(date_str)
    if date_str in holiday_reasons:
        del holiday_reasons[date_str]
    if date_str in holiday_messages:
        del holiday_messages[date_str]
    
    if status == "holiday":
        custom_holidays.add(date_str)
        holiday_reasons[date_str] = reason
        if message:
            holiday_messages[date_str] = message
    elif status == "working":
        custom_working_days.add(date_str)
        
    settings["custom_holidays"] = sorted(list(custom_holidays))
    settings["custom_working_days"] = sorted(list(custom_working_days))
    settings["holiday_reasons"] = holiday_reasons
    settings["holiday_messages"] = holiday_messages
    
    if save_settings(settings):
        role, u_name = get_log_identity()
        log_activity(role, u_name, f"Set day status: {date_str} -> {status} ({reason})")
        return jsonify({"success": True, "message": "Holiday status updated."})
    return jsonify({"success": False, "message": "Failed to save settings."}), 500

# =========================================================
# AUTHENTICATION
# =========================================================
@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    role = data.get("role","")

    if role == "principal":
        s = load_settings()
        if (data.get("username") == s.get("principal_username","principal") and
                data.get("password") == s.get("principal_password","principal123")):
            return jsonify({"success": True, "role": "principal",
                            "name": s.get("principal_name","Dr. Principal")})
        return jsonify({"success": False, "message": "Invalid credentials."})

    elif role == "teacher":
        phone = clean_phone(data.get("phone",""))
        pwd = data.get("password","").strip()
        for t in load_teachers():
            if clean_phone(t.get("phone","")) == phone:
                if not t.get("password",""):
                    return jsonify({"success": True, "role": "teacher",
                                    "teacher_id": t["id"], "name": t["name"],
                                    "require_password_setup": True})
                if t["password"] == pwd:
                    return jsonify({"success": True, "role": "teacher",
                                    "teacher_id": t["id"], "name": t["name"],
                                    "require_password_setup": False})
                return jsonify({"success": False, "message": "Incorrect password."})
        return jsonify({"success": False, "message": "Phone number not registered."})

    elif role == "parent":
        roll = data.get("roll_no","").strip()
        phone = data.get("phone_no","").strip()
        pwd = data.get("password","").strip()
        for s in load_db():
            if s["roll_no"] == roll or s["id"] == roll:
                if (clean_phone(s.get("parent_contact","")) == clean_phone(phone) or
                        clean_phone(s.get("parent_alt_contact","")) == clean_phone(phone)):
                    saved = s.get("parent_password","")
                    if saved and saved != pwd:
                        return jsonify({"success": False, "message": "Incorrect password."})
                    return jsonify({"success": True, "role": "parent", "student_id": s["id"],
                                    "name": s["parent_name"],
                                    "require_password_setup": not bool(saved)})
        return jsonify({"success": False, "message": "Roll number and phone do not match."})

    elif role == "admin":
        s = load_settings()
        if (data.get("username") == s.get("admin_username","admin") and
                data.get("password") == s.get("admin_password","admin123")):
            return jsonify({"success": True, "role": "admin",
                            "name": s.get("admin_name","School Administrator")})
        return jsonify({"success": False, "message": "Invalid credentials."})

    return jsonify({"success": False, "message": "Invalid role."})

# =========================================================
# OTP — Forgot Password & Activity Log
# =========================================================
def get_log_identity():
    role = request.headers.get('X-User-Role', 'Unknown')
    name = request.headers.get('X-User-Name', 'Unknown')
    return role.capitalize(), name

@app.route('/api/activities', methods=['GET'])
def get_activities_log():
    return jsonify(load_activities()[:100])

@app.route('/api/admin/reset-password', methods=['POST'])
def admin_reset_password_endpoint():
    data = request.get_json() or {}
    entity_id = data.get("id", "")
    entity_type = data.get("type", "")
    new_pwd = data.get("password", "").strip()
    
    if not entity_id or not entity_type or not new_pwd:
        return jsonify({"success": False, "message": "Missing required fields"}), 400
        
    role, name = get_log_identity()
    if entity_type == "student":
        students = load_db()
        for s in students:
            if s["id"] == entity_id:
                s["parent_password"] = new_pwd
                save_db(students)
                log_activity(role, name, f"Reset password for Student/Parent: {s['name']} (Class {s['class']}-{s['section']})")
                return jsonify({"success": True, "message": f"Password reset successful for {s['name']}."})
        return jsonify({"success": False, "message": "Student not found"}), 404
        
    elif entity_type == "teacher":
        teachers = load_teachers()
        for t in teachers:
            if t["id"] == entity_id:
                t["password"] = new_pwd
                save_teachers(teachers)
                log_activity(role, name, f"Reset password for Teacher: {t['name']}")
                return jsonify({"success": True, "message": f"Password reset successful for {t['name']}."})
        return jsonify({"success": False, "message": "Teacher not found"}), 404
        
    return jsonify({"success": False, "message": "Invalid entity type"}), 400

@app.route('/api/otp/send', methods=['POST'])
def send_otp():
    data = request.get_json() or {}
    phone = clean_phone(data.get("phone",""))
    role = data.get("role","")
    if not phone:
        return jsonify({"success": False, "message": "Phone number required."}), 400

    settings = load_settings()
    principal_phone = clean_phone(settings.get("principal_phone", "9999999999"))

    found = False
    if role == "teacher":
        found = any(clean_phone(t.get("phone","")) == phone for t in load_teachers())
    elif role == "parent":
        for s in load_db():
            if (clean_phone(s.get("parent_contact","")) == phone or
                    clean_phone(s.get("parent_alt_contact","")) == phone):
                found = True; break
    elif role == "principal":
        found = (phone == principal_phone)

    if not found:
        return jsonify({"success": False, "message": "Phone number not registered or incorrect."}), 404

    otp = ''.join(random.choices(string.digits, k=6))
    OTP_STORE[phone] = otp
    return jsonify({"success": True, "otp_demo": otp,
                    "message": f"OTP sent to {phone[-4:].rjust(10,'*')}"})

@app.route('/api/otp/verify', methods=['POST'])
def verify_otp():
    data = request.get_json() or {}
    phone = clean_phone(data.get("phone",""))
    otp = data.get("otp","").strip()
    if OTP_STORE.get(phone) == otp:
        return jsonify({"success": True})
    return jsonify({"success": False, "message": "Invalid or expired OTP."}), 400

@app.route('/api/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json() or {}
    phone = clean_phone(data.get("phone",""))
    new_pwd = data.get("password","").strip()
    role = data.get("role","")
    if not new_pwd:
        return jsonify({"success": False, "message": "Password cannot be empty."}), 400

    if role == "teacher":
        teachers = load_teachers()
        for t in teachers:
            if clean_phone(t.get("phone","")) == phone:
                t["password"] = new_pwd
                save_teachers(teachers)
                OTP_STORE.pop(phone, None)
                log_activity("Teacher", t["name"], "Reset password using OTP Forgot Password flow")
                return jsonify({"success": True})
    elif role == "parent":
        students = load_db()
        for s in students:
            if (clean_phone(s.get("parent_contact","")) == phone or
                    clean_phone(s.get("parent_alt_contact","")) == phone):
                s["parent_password"] = new_pwd
                save_db(students)
                OTP_STORE.pop(phone, None)
                log_activity("Parent", s["parent_name"], f"Reset password using OTP Forgot Password flow for {s['name']}")
                return jsonify({"success": True})
    elif role == "principal":
        settings = load_settings()
        settings["principal_password"] = new_pwd
        save_settings(settings)
        OTP_STORE.pop(phone, None)
        log_activity("Principal", settings.get("principal_name", "Principal"), "Reset password using OTP Forgot Password flow")
        return jsonify({"success": True})

    return jsonify({"success": False, "message": "Reset failed."}), 400

# =========================================================
# PASSWORD SETUP
# =========================================================
@app.route('/api/teacher/set-password', methods=['POST'])
def teacher_set_password():
    data = request.get_json() or {}
    tid = data.get("teacher_id","")
    pwd = data.get("password","").strip()
    teachers = load_teachers()
    for t in teachers:
        if t["id"] == tid:
            t["password"] = pwd
            save_teachers(teachers)
            return jsonify({"success": True})
    return jsonify({"success": False}), 404

@app.route('/api/parent/set-password', methods=['POST'])
def parent_set_password():
    data = request.get_json() or {}
    sid = data.get("student_id","")
    pwd = data.get("password","").strip()
    students = load_db()
    for s in students:
        if s["id"] == sid:
            s["parent_password"] = pwd
            save_db(students)
            return jsonify({"success": True})
    return jsonify({"success": False}), 404

@app.route('/api/parent/feedback', methods=['POST'])
def parent_feedback():
    data = request.get_json() or {}
    sid = data.get("student_id","")
    fb = data.get("feedback","").strip()
    students = load_db()
    for s in students:
        if s["id"] == sid:
            s["parent_feedback"] = fb
            save_db(students)
            return jsonify({"success": True})
    return jsonify({"success": False}), 404

@app.route('/api/principal/reply', methods=['POST'])
def principal_reply():
    data = request.get_json() or {}
    sid = data.get("student_id","")
    reply = data.get("reply","").strip()
    students = load_db()
    for s in students:
        if s["id"] == sid:
            s["principal_reply"] = reply
            save_db(students)
            return jsonify({"success": True})
    return jsonify({"success": False}), 404

# =========================================================
# TEACHERS CRUD
# =========================================================
@app.route('/api/teachers', methods=['GET'])
def get_teachers():
    teachers = load_teachers()
    res = []
    for t in teachers:
        classes = t.get("classes")
        if not classes and t.get("class") and t.get("section"):
            classes = [f"{t['class']}-{t['section']}"]
        elif not classes:
            classes = []
            
        res.append({
            "id": t["id"], "name": t["name"], "phone": t["phone"],
            "subjects": t.get("subjects",[]), "classes": classes,
            "attendance_status": t.get("attendance_status","Present"),
            "email": t.get("email",""), "has_password": bool(t.get("password","")),
            "can_edit_timetable": t.get("can_edit_timetable", False)
        })
    return jsonify(res)

@app.route('/api/teacher/create', methods=['POST'])
def create_teacher():
    import time
    data = request.get_json() or {}
    phone = clean_phone(data.get("phone",""))
    name = data.get("name","").strip()
    if not phone or not name:
        return jsonify({"success": False, "message": "Name and Phone required."}), 400
    if any(clean_phone(t.get("phone","")) == phone for t in load_teachers()):
        return jsonify({"success": False, "message": "Teacher with this phone already exists."}), 400
    teachers = load_teachers()
    new_t = {
        "id": f"t{int(time.time())}",
        "name": name, "phone": phone, "password": "",
        "subjects": data.get("subjects",[]),
        "classes": data.get("classes",[]),
        "attendance_status": "Present",
        "email": data.get("email",""),
        "can_edit_timetable": data.get("can_edit_timetable", False)
    }
    teachers.append(new_t)
    if save_teachers(teachers):
        role, u_name = get_log_identity()
        log_activity(role, u_name, f"Added teacher staff profile: {name}")
        return jsonify({"success": True, "teacher_id": new_t["id"]})
    return jsonify({"success": False}), 500

@app.route('/api/teacher/update/<tid>', methods=['POST'])
def update_teacher(tid):
    data = request.get_json() or {}
    teachers = load_teachers()
    for t in teachers:
        if t["id"] == tid:
            role, u_name = get_log_identity()
            changes = []
            
            if data.get("name") and t["name"] != data["name"].strip():
                changes.append(f"name ({t['name']} -> {data['name'].strip()})")
                t["name"] = data["name"].strip()
                
            new_phone = clean_phone(data.get("phone", ""))
            if new_phone and t["phone"] != new_phone:
                changes.append(f"phone ({t['phone']} -> {new_phone})")
                t["phone"] = new_phone
                
            if "subjects" in data:
                old_subs = t.get("subjects", [])
                new_subs = data["subjects"]
                if old_subs != new_subs:
                    changes.append(f"subjects ({', '.join(old_subs)} -> {', '.join(new_subs)})")
                    t["subjects"] = new_subs
                    
            if "classes" in data:
                old_classes = t.get("classes", [])
                new_classes = data["classes"]
                if old_classes != new_classes:
                    changes.append(f"classes assigned ({', '.join(old_classes)} -> {', '.join(new_classes)})")
                    t["classes"] = new_classes
                    
            if "email" in data and t.get("email") != data["email"]:
                changes.append(f"email ({t.get('email')} -> {data['email']})")
                t["email"] = data["email"]
                
            if "attendance_status" in data and t.get("attendance_status") != data["attendance_status"]:
                changes.append(f"attendance ({t.get('attendance_status')} -> {data['attendance_status']})")
                t["attendance_status"] = data["attendance_status"]
                
            new_edit_tt = data.get("can_edit_timetable", False)
            if t.get("can_edit_timetable", False) != new_edit_tt:
                changes.append(f"can_edit_timetable ({t.get('can_edit_timetable')} -> {new_edit_tt})")
                t["can_edit_timetable"] = new_edit_tt
                
            if save_teachers(teachers):
                desc = f"Edited Teacher {t['name']}"
                if changes:
                    desc += f": {', '.join(changes)}"
                log_activity(role, u_name, desc)
                return jsonify({"success": True})
            return jsonify({"success": False}), 500
    return jsonify({"success": False}), 404

@app.route('/api/teacher/delete/<tid>', methods=['POST', 'DELETE'])
def delete_teacher(tid):
    teachers = load_teachers()
    t_obj = next((x for x in teachers if x["id"] == tid), None)
    new_list = [t for t in teachers if t["id"] != tid]
    if len(new_list) == len(teachers): return jsonify({"success": False}), 404
    if save_teachers(new_list):
        if t_obj:
            role, u_name = get_log_identity()
            log_activity(role, u_name, f"Removed teacher staff profile: {t_obj['name']}")
        return jsonify({"success": True})
    return jsonify({"success": False}), 500

@app.route('/api/teacher/attendance', methods=['POST'])
def update_teacher_attendance():
    data = request.get_json() or {}
    teachers = load_teachers()
    for t in teachers:
        if t["id"] == data.get("teacher_id",""):
            old_status = t.get("attendance_status", "Present")
            new_status = data.get("status","Present")
            t["attendance_status"] = new_status
            if save_teachers(teachers):
                role, u_name = get_log_identity()
                log_activity(role, u_name, f"Updated Teacher Attendance for {t['name']}: {old_status} -> {new_status}")
                return jsonify({"success": True})
            return jsonify({"success": False}), 500
    return jsonify({"success": False}), 404

# =========================================================
# TIMETABLE
# =========================================================
@app.route('/api/timetable/<class_key>', methods=['GET'])
def get_timetable(class_key):
    timetables = load_timetables()
    return jsonify({"success": True, "timetable": timetables.get(class_key)})

@app.route('/api/timetable/save', methods=['POST'])
def save_timetable():
    data = request.get_json() or {}
    class_key = data.get("class_key","")
    timetable = data.get("timetable",{})
    if not class_key: return jsonify({"success": False}), 400
    timetables = load_timetables()
    timetables[class_key] = timetable
    if save_timetables(timetables):
        role, u_name = get_log_identity()
        log_activity(role, u_name, f"Updated Timetable for Class {class_key}")
        # Propagate to students in this class-section
        parts = class_key.split("-",1)
        if len(parts) == 2:
            students = load_db()
            for s in students:
                if str(s["class"]) == parts[0] and s["section"] == parts[1]:
                    s["timetable"] = timetable
            save_db(students)
        return jsonify({"success": True})
    return jsonify({"success": False}), 500

@app.route('/api/timetable/subjects/<class_key>', methods=['GET'])
def get_timetable_subjects(class_key):
    parts = class_key.split("-", 1)
    if len(parts) != 2:
        return jsonify({"success": False, "message": "Invalid class key"}), 400
    subs = get_subjects_from_timetable(parts[0], parts[1])
    return jsonify({"success": True, "subjects": subs or []})

@app.route('/api/exams/list', methods=['GET'])
def list_published_exams():
    students = load_db()
    exams = set()
    for s in students:
        for e in s.get('examination_progress', []):
            name = e.get('exam_name') or e.get('exam')
            if name:
                exams.add(name)
    # Fallback to standard exams if empty
    standard = ['Unit Test 1', 'Unit Test 2', 'Quarterly', 'Half-Yearly', 'Pre-Final', 'Final Examination']
    all_exams = sorted(list(exams)) if exams else standard
    return jsonify({"success": True, "exams": all_exams})

# =========================================================
# ATTENDANCE
# =========================================================
@app.route('/api/attendance/save', methods=['POST'])
def save_attendance():
    data = request.get_json() or {}
    date = data.get("date","")
    records = data.get("records",[])
    students = load_db()
    
    role, name = get_log_identity()
    changed_records = []
    target_class_sec = "Unknown"
    
    for rec in records:
        for s in students:
            if s["id"] == rec["student_id"]:
                if target_class_sec == "Unknown":
                    target_class_sec = f"Class {s.get('class','')}-{s.get('section','')}"
                
                if "attendance_records" not in s: s["attendance_records"] = []
                existing = next((r for r in s["attendance_records"] if r["date"]==date), None)
                
                old_status = existing["status"] if existing else "None"
                new_status = rec["status"]
                
                if old_status != new_status:
                    changed_records.append(f"{s['name']} ({old_status} -> {new_status})")
                
                if existing: existing["status"] = rec["status"]
                else: s["attendance_records"].append({"date": date, "status": rec["status"]})
                s["attendance_status"] = "Absent" if rec["status"]=="Absent" else "Present"
                
                # Recalculate attendance percentage
                recs = s.get('attendance_records', [])
                workdays = sum(1 for item in recs if item.get('status') != 'Holiday')
                present_days = sum(1 for item in recs if item.get('status') == 'Present')
                half_days = sum(1 for item in recs if item.get('status') == 'Half Day')
                effective_present = present_days + (half_days / 2.0)
                
                if workdays > 0:
                    s['current_status']['attendance_percentage'] = round((effective_present / workdays) * 100, 1)
                else:
                    s['current_status']['attendance_percentage'] = 100.0

    if save_db(students):
        if changed_records:
            desc = f"Saved Attendance for {target_class_sec} on {date}: {', '.join(changed_records)}"
            log_activity(role, name, desc)
        return jsonify({"success": True, "updated": len(records)})
    return jsonify({"success": False}), 500

@app.route('/api/attendance/class', methods=['GET'])
def get_class_attendance():
    cls = request.args.get('class','')
    sec = request.args.get('section','')
    date = request.args.get('date','')
    students = load_db()
    result = []
    for s in students:
        if cls and str(s["class"]) != cls: continue
        if sec and s["section"] != sec: continue
        status = "Not Marked"
        if date:
            rec = next((r for r in s.get("attendance_records",[]) if r["date"]==date), None)
            if rec: status = rec["status"]
        result.append({"id": s["id"], "name": s["name"], "roll_no": s["roll_no"],
                        "attendance_status": status})
    return jsonify(sorted(result, key=lambda x: x["roll_no"]))

@app.route('/api/attendance/report', methods=['GET'])
def attendance_report():
    cls = request.args.get('class','')
    sec = request.args.get('section','')
    month = request.args.get('month','')
    students = load_db()
    report = []
    for s in students:
        if cls and str(s["class"]) != cls: continue
        if sec and s["section"] != sec: continue
        recs = [r for r in s.get("attendance_records",[]) if (not month or r["date"].startswith(month))]
        workdays = sum(1 for r in recs if r["status"] != "Holiday")
        present = sum(1 for r in recs if r["status"] == "Present")
        absent = sum(1 for r in recs if r["status"] == "Absent")
        half = sum(1 for r in recs if r["status"] == "Half Day")
        effective_present = present + (half / 2.0)
        percentage = round((effective_present / workdays) * 100, 1) if workdays > 0 else 0.0
        report.append({"id": s["id"], "name": s["name"], "roll_no": s["roll_no"],
                        "class": s["class"], "section": s["section"],
                        "present_days": present, "absent_days": absent, "half_days": half,
                        "total_days": workdays,
                        "percentage": percentage})
    return jsonify(report)

# =========================================================
# PARENT MEETINGS (Principal manages)
# =========================================================
@app.route('/api/meetings', methods=['GET'])
def get_meetings():
    return jsonify(load_meetings())

@app.route('/api/meetings/save', methods=['POST'])
def save_meeting():
    import time
    data = request.get_json() or {}
    meetings = load_meetings()
    meeting_id = data.get("id","")
    role, name = get_log_identity()
    if meeting_id:
        for m in meetings:
            if m["id"] == meeting_id:
                m.update({k: data[k] for k in ["title","date","time","venue","classes","notes"] if k in data})
                if save_meetings(meetings):
                    log_activity(role, name, f"Edited parent meeting: '{m.get('title')}' on {m.get('date')}")
                    return jsonify({"success": True})
                return jsonify({"success": False}), 500
    else:
        new_m = {
            "id": f"mtg{int(time.time())}",
            "title": data.get("title","Parent Meeting"),
            "date": data.get("date",""),
            "time": data.get("time",""),
            "venue": data.get("venue",""),
            "classes": data.get("classes","all"),  # "all" or list of class-section strings
            "notes": data.get("notes","")
        }
        meetings.append(new_m)
        if save_meetings(meetings):
            log_activity(role, name, f"Scheduled parent meeting: '{new_m['title']}' on {new_m['date']}")
            return jsonify({"success": True, "id": new_m["id"]})
        return jsonify({"success": False}), 500

@app.route('/api/meetings/delete/<mid>', methods=['POST','DELETE'])
def delete_meeting(mid):
    meetings = load_meetings()
    deleted_title = "Meeting"
    for m in meetings:
        if m["id"] == mid:
            deleted_title = m.get("title", "Meeting")
            break
    new_list = [m for m in meetings if m["id"] != mid]
    if save_meetings(new_list):
        role, name = get_log_identity()
        log_activity(role, name, f"Deleted parent meeting: '{deleted_title}'")
        return jsonify({"success": True})
    return jsonify({"success": False}), 500

# =========================================================
# STUDENT CRUD
# =========================================================
@app.route('/api/admin/student/create', methods=['POST'])
def admin_create_student():
    import time
    data = request.get_json() or {}
    parent_contact = clean_phone(data.get("parent_contact",""))
    if not parent_contact:
        return jsonify({"success": False, "message": "Parent contact is required!"}), 400
    name = data.get("name","").strip()
    roll_no = data.get("roll_no","").strip()
    if not name or not roll_no:
        return jsonify({"success": False, "message": "Name and Roll Number required!"}), 400

    students = load_db()
    s_class = str(data.get("class","10"))
    s_section = data.get("section","A")
    if any(s["class"]==s_class and s["section"]==s_section and s["roll_no"]==roll_no for s in students):
        return jsonify({"success": False, "message": f"Roll {roll_no} already in {s_class}-{s_section}"}), 400

    class_key = f"{s_class}-{s_section}"
    timetables = load_timetables()
    class_timetable = timetables.get(class_key, {})

    student_id = str(int(time.time()))
    new_student = {
        "id": student_id,
        "name": name, "roll_no": roll_no,
        "admission_no": data.get("admission_no",""),
        "class": s_class, "section": s_section,
        "academic_year": data.get("academic_year","2025-26"),
        "dob": data.get("dob",""),
        "parent_name": data.get("parent_name",""),
        "parent_contact": parent_contact,
        "parent_alt_contact": clean_phone(data.get("parent_alt_contact","")),
        "parent_password": "", "parent_feedback": "", "principal_reply": "",
        "attendance_status": data.get("attendance_status","Present"),
        "attendance_records": [],
        "fees": {
            "school": 10000,
            "tuition": 30000,
            "books": 5000,
            "dresses": 3000,
            "extra": 0,
            "paid": 0
        },
        "current_status": {
            "standard": f"Class {s_class}", "section": f"Section {s_section}",
            "class_teacher": data.get("class_teacher","Class Teacher"),
            "attendance_percentage": float(data.get("attendance_percentage",90)),
            "subjects": data.get("subjects",["Mathematics","Science","English","Hindi","Social Studies"])
        },
        "timetable": class_timetable,
        "examination_progress": data.get("examination_progress",[]),
        "subject_performance": data.get("subject_performance",[]),
        "progress_comparison": data.get("progress_comparison",[]),
        "performance_trend": data.get("performance_trend","Stable"),
        "teacher_term_remarks": data.get("teacher_term_remarks",[]),
        "behavioral_observation": data.get("behavioral_observation",
            {"discipline":4,"leadership":4,"participation":4,"communication":4,"teamwork":4,"confidence":4}),
        "co_curricular_activities": data.get("co_curricular_activities",[]),
        "awards": data.get("awards",[]),
        "parent_meetings": data.get("parent_meetings",[])
    }
    students.append(new_student)
    if save_db(students):
        user_role, user_name = get_log_identity()
        log_activity(user_role, user_name, f"Registered new Student: {name} (Class {s_class}-{s_section})")
        return jsonify({"success": True, "student_id": student_id})
    return jsonify({"success": False}), 500

@app.route('/api/admin/student/update/<student_id>', methods=['POST'])
def admin_update_student(student_id):
    data = request.get_json() or {}
    students = load_db()
    for s in students:
        if s["id"] == student_id:
            user_role, user_name = get_log_identity()
            changes = []
            
            # Simple fields
            fields_to_check = ["name", "roll_no", "admission_no", "dob", "academic_year", "parent_name", "performance_trend", "attendance_status"]
            for field in fields_to_check:
                if field in data:
                    old_val = s.get(field, "")
                    new_val = data[field]
                    if str(old_val) != str(new_val):
                        changes.append(f"{field} ({old_val} -> {new_val})")
                        s[field] = new_val

            # Class / section
            new_class = str(data.get("class", s["class"]))
            if str(s["class"]) != new_class:
                changes.append(f"class ({s['class']} -> {new_class})")
                s["class"] = new_class
                
            new_section = data.get("section", s["section"])
            if s["section"] != new_section:
                changes.append(f"section ({s['section']} -> {new_section})")
                s["section"] = new_section

            # Contacts
            pc = clean_phone(data.get("parent_contact",""))
            if not pc: return jsonify({"success": False, "message": "Parent contact required!"}), 400
            if s.get("parent_contact") != pc:
                changes.append(f"parent_contact ({s.get('parent_contact')} -> {pc})")
                s["parent_contact"] = pc
                
            alt = clean_phone(data.get("parent_alt_contact",""))
            if s.get("parent_alt_contact") != alt:
                changes.append(f"parent_alt_contact ({s.get('parent_alt_contact')} -> {alt})")
                s["parent_alt_contact"] = alt

            # Check exam marks updates
            if "examination_progress" in data:
                old_progress = s.get("examination_progress", [])
                new_progress = data["examination_progress"]
                for new_ex in new_progress:
                    exam_name = new_ex.get("exam_name") or new_ex.get("exam")
                    old_ex = next((e for e in old_progress if (e.get("exam_name") or e.get("exam")) == exam_name), None)
                    if not old_ex:
                        changes.append(f"added marks for exam '{exam_name}'")
                    else:
                        sub_changes = []
                        for sub in new_ex.get("subjects", []):
                            old_sub = next((sb for sb in old_ex.get("subjects", []) if sb.get("name") == sub.get("name")), None)
                            if not old_sub or old_sub.get("obtained") != sub.get("obtained"):
                                old_obt = old_sub.get("obtained") if old_sub else "None"
                                sub_changes.append(f"{sub.get('name')}: {old_obt} -> {sub.get('obtained')}")
                        if sub_changes:
                            changes.append(f"updated '{exam_name}' marks: {', '.join(sub_changes)}")

            # Principal reply
            old_pwd = s.get("parent_password","")
            old_fb = s.get("parent_feedback","")
            old_reply = s.get("principal_reply","")
            
            new_reply = data.get("principal_reply", old_reply)
            if old_reply != new_reply:
                changes.append(f"principal_reply ({old_reply} -> {new_reply})")
                s["principal_reply"] = new_reply
            else:
                s["principal_reply"] = old_reply

            if "current_status" not in s: s["current_status"] = {}
            s["current_status"]["standard"] = f"Class {s['class']}"
            s["current_status"]["section"] = f"Section {s['section']}"
            if "class_teacher" in data: s["current_status"]["class_teacher"] = data["class_teacher"]
            if "attendance_percentage" in data: s["current_status"]["attendance_percentage"] = float(data["attendance_percentage"])
            if "subjects" in data: s["current_status"]["subjects"] = data["subjects"]

            for field in ["timetable","examination_progress","subject_performance","progress_comparison",
                          "teacher_term_remarks","behavioral_observation","co_curricular_activities","awards","parent_meetings"]:
                if field in data: s[field] = data[field]

            s["parent_password"] = old_pwd
            s["parent_feedback"] = old_fb

            if save_db(students):
                desc = f"Updated Student: {s['name']} (Class {s['class']}-{s['section']})"
                if changes:
                    desc += f": {', '.join(changes)}"
                log_activity(user_role, user_name, desc)
                return jsonify({"success": True})
            return jsonify({"success": False}), 500
    return jsonify({"success": False, "message": "Not found"}), 404

@app.route('/api/admin/student/delete/<student_id>', methods=['POST','DELETE'])
def admin_delete_student(student_id):
    students = load_db()
    deleted_student = None
    for s in students:
        if s["id"] == student_id:
            deleted_student = s
            break
    if not deleted_student: return jsonify({"success": False}), 404
    new_s = [s for s in students if s["id"] != student_id]
    if save_db(new_s):
        user_role, user_name = get_log_identity()
        log_activity(user_role, user_name, f"Deleted Student: {deleted_student['name']} (Class {deleted_student.get('class','')}-{deleted_student.get('section','')})")
        return jsonify({"success": True})
    return jsonify({"success": False}), 500

def get_teacher_classes(teacher_id):
    if not teacher_id:
        return None
    for t in load_teachers():
        if t["id"] == teacher_id:
            classes = t.get("classes")
            if not classes and t.get("class") and t.get("section"):
                return [f"{t['class']}-{t['section']}"]
            return classes or []
    return []

@app.route('/api/search', methods=['GET'])
def search_students():
    q = request.args.get('q','').lower()
    cls = request.args.get('class','')
    sec = request.args.get('section','')
    teacher_id = request.headers.get('X-Teacher-Id', '')
    allowed_classes = get_teacher_classes(teacher_id)
    
    students = load_db()
    def summary(s):
        return {"id": s["id"], "name": s["name"], "admission_no": s.get("admission_no",""),
                "roll_no": s["roll_no"], "class": s["class"], "section": s["section"],
                "parent_name": s.get("parent_name",""), "parent_contact": s.get("parent_contact",""),
                "attendance": s["current_status"]["attendance_percentage"],
                "trend": s.get("performance_trend","Stable"),
                "parent_feedback": s.get("parent_feedback",""),
                "principal_reply": s.get("principal_reply",""),
                "attendance_status": s.get("attendance_status","Present")}
    results = []
    for s in students:
        s_key = f"{s['class']}-{s['section']}"
        if allowed_classes is not None and s_key not in allowed_classes:
            continue
        if cls and str(s["class"]) != cls: continue
        if sec and s["section"] != sec: continue
        if q and not any([q in s["name"].lower(), q in s.get("admission_no","").lower(),
                          q == s["roll_no"], q in s.get("parent_contact",""),
                          q in s.get("parent_alt_contact","")]): continue
        results.append(summary(s))
    return jsonify(results)

@app.route('/api/student/<student_id>', methods=['GET'])
def get_student(student_id):
    teacher_id = request.headers.get('X-Teacher-Id', '')
    allowed_classes = get_teacher_classes(teacher_id)
    
    for s in load_db():
        if s["id"] == student_id:
            s_key = f"{s['class']}-{s['section']}"
            if allowed_classes is not None and s_key not in allowed_classes:
                return jsonify({"error": "Unauthorized"}), 403
            return jsonify(s)
    return jsonify({"error": "Not found"}), 404

# =========================================================
# EXCEL HELPERS
# =========================================================
DEFAULT_SUBJECTS = ['Mathematics', 'Science', 'English', 'Hindi', 'Social Studies', 'Computer Science']
CLASS_LIST = ['0', 'Nursery', 'LKG', 'UKG', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10']
EXAM_LIST = ['Unit Test 1', 'Unit Test 2', 'Quarterly', 'Half-Yearly', 'Pre-Final', 'Final Examination']

def get_grade(pct):
    if pct >= 90: return 'A+'
    if pct >= 80: return 'A'
    if pct >= 70: return 'B+'
    if pct >= 60: return 'B'
    if pct >= 50: return 'C'
    if pct >= 35: return 'D'
    return 'F'


def style_header(ws, row, cols, fill_hex='1E3A5F', font_size=11):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type='solid')
    font = Font(bold=True, color='FFFFFF', size=font_size)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='999999')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font
        cell.alignment = align; cell.border = border

def style_data_row(ws, row, cols, even=False):
    fill = PatternFill(start_color='F0F4FF' if even else 'FFFFFF',
                       end_color='F0F4FF' if even else 'FFFFFF', fill_type='solid')
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    align = Alignment(horizontal='center', vertical='center')
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.border = border; cell.alignment = align

# =========================================================
# DOWNLOAD SAMPLE EXCEL
# =========================================================
def get_subjects_from_timetable(cls, sec):
    class_key = f"{cls}-{sec}"
    timetables = load_timetables()
    timetable = timetables.get(class_key, {})
    subjects = set()
    non_subjects = ["free", "lunch", "break", "recess", "interval", "games", "sports", "pt", "lib", "library", "breakfast", "play", "assembly", "prayer", "leisure", "zero", "time", "period", "hour", "gap", "snack", "snacks", "tea", "recreation", "club", "pe", "moral", "value", "values", "activity", "activities", "house", "test", "exam", "revision", "gk", "general knowledge", "drawing", "art", "craft", "music", "dance", "games/play", "drill"]
    for day, periods in timetable.items():
        for p in periods:
            status = p.get("status", "Class In Progress")
            if status != "Class In Progress":
                continue
            sub = p.get("subject", "").strip()
            sub_lower = sub.lower()
            if sub and not any(ns in sub_lower for ns in non_subjects):
                subjects.add(sub)
    return sorted(list(subjects)) if subjects else None

# =========================================================
# DOWNLOAD SAMPLE EXCEL
# =========================================================
@app.route('/api/excel/sample-marks', methods=['GET'])
def download_sample_marks():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500

    cls = request.args.get('class', '10')
    sec = request.args.get('section', 'A')
    exam = request.args.get('exam', 'Unit Test 1').strip()
    subjects_str = request.args.get('subjects', '')
    
    if subjects_str:
        subjects = [s.strip() for s in subjects_str.split(',') if s.strip()]
    else:
        subjects = ['Telugu', 'Hindi', 'English', 'Mathematics', 'Science', 'Social Studies']

    students = load_db()
    class_students = sorted(
        [s for s in students if str(s['class']) == cls and s['section'] == sec],
        key=lambda x: x['roll_no']
    )

    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    subj_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    lock_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
    calc_fill  = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')

    ws = wb.create_sheet(title=exam.replace(' ', '_').replace('/', '_'))
    ws.freeze_panes = 'E3'

    # Title row
    total_cols = 4 + len(subjects) + 5  # roll,name,class,sec + subjects + total,pct,grade,rank,remarks
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f'SPTAS — {exam} Marks | Class {cls}-{sec} | Max/Subject: 100'
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    # Header row
    headers = ['Roll No', 'Student Name', 'Class', 'Section'] + subjects + \
              ['Total', 'Percentage', 'Grade', 'Class Rank', 'Remarks']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col).value = h
    style_header(ws, 2, len(headers))

    # Colour subject columns
    for col in range(5, 5 + len(subjects)):
        ws.cell(row=2, column=col).fill = PatternFill(start_color='1D6A3A', end_color='1D6A3A', fill_type='solid')
        ws.cell(row=2, column=col).font = Font(bold=True, color='FFFFFF', size=11)

    # Colour calculated columns
    calc_start = 5 + len(subjects)
    for col in range(calc_start, calc_start + 4):
        ws.cell(row=2, column=col).fill = PatternFill(start_color='0C4A6E', end_color='0C4A6E', fill_type='solid')
        ws.cell(row=2, column=col).font = Font(bold=True, color='FFFFFF', size=11)

    # Student rows
    for row_i, student in enumerate(class_students, start=3):
        row_data = [
            student['roll_no'], student['name'],
            student['class'], student['section']
        ] + [0] * len(subjects) + ['', '', '', '', '']
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            if col <= 4:
                cell.fill = lock_fill
                cell.font = Font(size=10, bold=(col==1))
            elif col <= 4 + len(subjects):
                cell.fill = subj_fill
                cell.font = Font(size=11)
            else:
                cell.fill = calc_fill
                cell.font = Font(size=10, color='555555')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CCCCCC')
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # If no students, add example rows
    if not class_students:
        for ex_i, ex_name in enumerate(['Example Student 1', 'Example Student 2'], start=3):
            row_data = [ex_i - 2, ex_name, cls, sec] + [0]*len(subjects) + ['','','','',' ']
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=ex_i, column=col, value=val)
                if col <= 4:
                    cell.fill = lock_fill
                elif col <= 4 + len(subjects):
                    cell.fill = subj_fill
                else:
                    cell.fill = calc_fill
                thin = Side(style='thin', color='CCCCCC')
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 10
    for c in range(5, 5 + len(subjects)):
        ws.column_dimensions[get_column_letter(c)].width = 14
    for c in range(5+len(subjects), 5+len(subjects)+5):
        ws.column_dimensions[get_column_letter(c)].width = 13

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'SPTAS_Marks_Template_{exam.replace(" ","_")}_Class{cls}{sec}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =========================================================
# DOWNLOAD STUDENT ATTENDANCE SAMPLE TEMPLATE
# =========================================================
@app.route('/api/excel/sample-attendance', methods=['GET'])
def download_sample_attendance():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500

    cls = request.args.get('class', '10')
    sec = request.args.get('section', 'A')

    students = load_db()
    class_students = sorted(
        [s for s in students if str(s['class']) == cls and s['section'] == sec],
        key=lambda x: x['roll_no']
    )

    wb = openpyxl.Workbook()
    att = wb.active
    att.title = 'Attendance'
    att.freeze_panes = 'E3'

    import calendar
    from datetime import date
    today = date.today()
    month_days = calendar.monthrange(today.year, today.month)[1]
    dates = [f'{today.year}-{today.month:02d}-{d:02d}' for d in range(1, month_days + 1)]

    # 4 (Roll, Name, Class, Sec) + len(dates) + 5 (Work Days, Present, Absent, Half Day, Attendance %)
    total_cols = 4 + len(dates) + 5
    att.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = att.cell(row=1, column=1)
    title_cell.value = f'SPTAS — Attendance Register | Class {cls}-{sec} | Month: {today.strftime("%B %Y")}'
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='064E3B', end_color='064E3B', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    att.row_dimensions[1].height = 28

    headers = ['Roll No', 'Student Name', 'Class', 'Section'] + dates + \
              ['Work Days', 'Present Days', 'Absent Days', 'Half Days', 'Attendance %']
    for col, h in enumerate(headers, 1):
        att.cell(row=2, column=col).value = h
    style_header(att, 2, len(headers), fill_hex='064E3B')

    # Color definitions
    lock_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
    holiday_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    holiday_font = Font(color='842029', bold=True, size=10)
    workday_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    calc_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    calc_font = Font(bold=True, color='065F46')

    # Prefill rows helper
    def write_student_row(row_idx, roll_val, name_val, class_val, sec_val):
        # 1. Basic details
        for col_idx, val in enumerate([roll_val, name_val, class_val, sec_val], 1):
            c = att.cell(row=row_idx, column=col_idx, value=val)
            c.fill = lock_fill
            c.alignment = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CCCCCC')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 2. Date columns
        for d_idx, date_str in enumerate(dates, start=5):
            is_hol, reason = is_school_holiday(date_str)
            c = att.cell(row=row_idx, column=d_idx)
            c.alignment = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CCCCCC')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if is_hol:
                c.value = 'Holiday'
                c.fill = holiday_fill
                c.font = holiday_font
            else:
                c.value = ''
                c.fill = workday_fill

        # 3. Dynamic Excel Formulas
        # Let's find column names
        n_dates = len(dates)
        col_work = get_column_letter(5 + n_dates)
        col_pres = get_column_letter(6 + n_dates)
        col_abs  = get_column_letter(7 + n_dates)
        col_half = get_column_letter(8 + n_dates)
        col_pct  = get_column_letter(9 + n_dates)

        # Formulas
        f_work = f'={n_dates}-COUNTIF(E{row_idx}:{get_column_letter(4+n_dates)}{row_idx},"Holiday")'
        f_pres = f'=COUNTIF(E{row_idx}:{get_column_letter(4+n_dates)}{row_idx},"P")'
        f_abs  = f'=COUNTIF(E{row_idx}:{get_column_letter(4+n_dates)}{row_idx},"A")'
        f_half = f'=COUNTIF(E{row_idx}:{get_column_letter(4+n_dates)}{row_idx},"HD")'
        f_pct  = f'=IF({col_work}{row_idx}>0,ROUND((({col_pres}{row_idx}+({col_half}{row_idx}/2))/{col_work}{row_idx})*100,2),0)'

        for c_idx, f_val in enumerate([f_work, f_pres, f_abs, f_half, f_pct], start=5+n_dates):
            c = att.cell(row=row_idx, column=c_idx, value=f_val)
            c.fill = calc_fill
            c.font = calc_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CCCCCC')
            c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write student rows
    for r_i, student in enumerate(class_students, start=3):
        write_student_row(r_i, student['roll_no'], student['name'], student['class'], student['section'])

    # Fallback to example rows if empty class
    if not class_students:
        for ex_i, ex_name in enumerate(['Example Student 1', 'Example Student 2'], start=3):
            write_student_row(ex_i, ex_i - 2, ex_name, cls, sec)

    # Apply column widths
    att.column_dimensions['A'].width = 10
    att.column_dimensions['B'].width = 24
    att.column_dimensions['C'].width = 8
    att.column_dimensions['D'].width = 10
    for c in range(5, 5 + len(dates)):
        att.column_dimensions[get_column_letter(c)].width = 12
    for c in range(5 + len(dates), 5 + len(dates) + 5):
        att.column_dimensions[get_column_letter(c)].width = 15

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'SPTAS_Attendance_Template_Class{cls}{sec}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =========================================================
# UPLOAD & PROCESS MARKS EXCEL
# =========================================================
@app.route('/api/excel/upload-marks', methods=['POST'])
def upload_marks_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'message': 'openpyxl not available'}), 500
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    f = request.files['file']
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Only .xlsx files allowed'}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Cannot read Excel: {str(e)}'}), 400

    students = load_db()
    student_map = {s['roll_no']: s for s in students}  # roll_no -> student
    results_summary = []
    updated_count = 0
    errors = []

    skip_sheets = {'Instructions', 'Grades & Keys', 'Attendance'}

    for sheet_name in wb.sheetnames:
        if sheet_name in skip_sheets:
            continue
        ws = wb[sheet_name]
        exam_name = sheet_name.replace('_', ' ')

        # Read header row (row 2)
        headers = []
        for col in range(1, ws.max_column + 1):
            h = ws.cell(row=2, column=col).value
            headers.append(str(h).strip() if h else '')

        if 'Roll No' not in headers:
            errors.append(f'Sheet "{sheet_name}": No "Roll No" column found, skipped.')
            continue

        roll_col = headers.index('Roll No') + 1
        name_col = headers.index('Student Name') + 1 if 'Student Name' in headers else None
        class_col = headers.index('Class') + 1 if 'Class' in headers else None
        sec_col = headers.index('Section') + 1 if 'Section' in headers else None

        # Identify subject columns (between Section and Total)
        subj_start = max(roll_col, name_col or 0, class_col or 0, sec_col or 0) + 1
        subj_cols = []
        calc_cols = {'Total', 'Percentage', 'Grade', 'Class Rank', 'Remarks'}
        for idx, h in enumerate(headers[subj_start - 1:], start=subj_start):
            if h in calc_cols or not h: break
            subj_cols.append((idx, h))

        # Find max_marks: assume 100 unless a note in title
        title_cell = ws.cell(row=1, column=1).value or ''
        max_marks = 100
        if 'Max/Subject:' in str(title_cell):
            try: max_marks = int(str(title_cell).split('Max/Subject:')[1].strip())
            except: pass

        sheet_rows = []
        for row in range(3, ws.max_row + 1):
            roll = str(ws.cell(row=row, column=roll_col).value or '').strip()
            if not roll or roll == 'None': continue

            subject_scores = {}
            for sc, sname in subj_cols:
                val = ws.cell(row=row, column=sc).value
                try: subject_scores[sname] = float(val or 0)
                except: subject_scores[sname] = 0

            total = sum(subject_scores.values())
            total_max = max_marks * len(subj_cols)
            pct = round((total / total_max) * 100, 2) if total_max > 0 else 0
            grade = get_grade(pct)

            sheet_rows.append({'roll_no': roll, 'subjects': subject_scores,
                                'total': total, 'total_max': total_max,
                                'percentage': pct, 'grade': grade})

        # Calculate class rank (by percentage desc)
        sheet_rows.sort(key=lambda x: -x['percentage'])
        for rank_i, row_data in enumerate(sheet_rows, start=1):
            row_data['rank'] = rank_i

        # Collect parsed rows with student names
        for row_data in sheet_rows:
            s = student_map.get(row_data['roll_no'])
            if s:
                row_data['name'] = s['name']
                row_data['class'] = s.get('class','')
                row_data['section'] = s.get('section','')
                row_data['exam'] = exam_name
                row_data['max_marks'] = max_marks
                results_summary.append(row_data)
            else:
                errors.append(f"Roll No '{row_data['roll_no']}' not found in database. Skipped.")

    return jsonify({
        'success': True,
        'preview': True,
        'type': 'marks',
        'data': results_summary,
        'errors': errors
    })



# =========================================================
# UPLOAD & PROCESS ATTENDANCE EXCEL
# =========================================================
@app.route('/api/excel/upload-attendance', methods=['POST'])
def upload_attendance_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'message': 'openpyxl not available'}), 500
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Cannot read file: {str(e)}'}), 400

    if 'Attendance' not in wb.sheetnames:
        return jsonify({'success': False, 'message': '"Attendance" sheet not found in file'}), 400

    ws = wb['Attendance']
    headers = [str(ws.cell(2, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    roll_col = 1  # always col 1
    date_cols = [(i+1, h) for i, h in enumerate(headers) if h.count('-') == 2 and len(h) == 10]

    students = load_db()
    student_map = {s['roll_no']: s for s in students}
    attendance_data = []
    for row in range(3, ws.max_row + 1):
        roll = str(ws.cell(row, roll_col).value or '').strip()
        if not roll or roll == 'None': continue
        s = student_map.get(roll)
        if not s: continue
        row_recs = []
        for col, date_str in date_cols:
            code = str(ws.cell(row, col).value or 'P').strip().upper()
            status = 'Present' if code == 'P' else ('Absent' if code == 'A' else 'Half Day')
            row_recs.append({'date': date_str, 'status': status})
        attendance_data.append({
            'roll_no': roll,
            'name': s['name'],
            'records': row_recs
        })

    return jsonify({
        'success': True,
        'preview': True,
        'type': 'attendance',
        'data': attendance_data,
        'date_columns': len(date_cols)
    })



# =========================================================
# DOWNLOAD RESULTS REPORT EXCEL
# =========================================================
@app.route('/api/excel/results-report', methods=['GET'])
def download_results_report():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500
    cls  = request.args.get('class', '')
    sec  = request.args.get('section', '')
    exam = request.args.get('exam', '')

    teacher_id = request.headers.get('X-Teacher-Id', '')
    allowed_classes = get_teacher_classes(teacher_id)
    if allowed_classes is not None:
        if cls and sec:
            s_key = f"{cls}-{sec}"
            if s_key not in allowed_classes:
                return jsonify({"error": "Forbidden: This is not your Class"}), 403
        elif cls:
            matched = [ac for ac in allowed_classes if ac.split('-')[0] == cls]
            if not matched:
                return jsonify({"error": "Forbidden: This is not your Class"}), 403
        else:
            return jsonify({"error": "Forbidden: Teacher must specify an assigned Class"}), 403

    students = load_db()
    filtered = sorted([s for s in students if (not cls or str(s['class'])==cls) and
                (not sec or s['section']==sec)], key=lambda x: x['roll_no'])

    # Gather subjects list from students' exam progress
    subjects = []
    for s in filtered:
        prog = s.get('examination_progress', [])
        exam_entry = next((e for e in prog if (e.get('exam_name') or e.get('exam','')) == exam), None)
        if exam_entry and exam_entry.get('subjects'):
            subjects = [sub.get('name') for sub in exam_entry['subjects']]
            break
            
    if not subjects:
        if cls:
            subjects = get_subjects_from_timetable(cls, sec or 'A')
        if not subjects:
            subjects = ['Telugu', 'Hindi', 'English', 'Mathematics', 'Science', 'Social Studies']

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Results Report'
    ws.freeze_panes = 'E3'

    # Title
    total_cols = 4 + len(subjects) + 6
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws['A1'] = f'SPTAS — Results Report | Class {cls or "All"}-{sec or "All"} | {exam or "All Exams"}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1E1B4B', end_color='1E1B4B', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = ['Roll No', 'Student Name', 'Class', 'Section'] + subjects + \
              ['Total', 'Percentage', 'Grade', 'Class Rank', 'Result', 'Remarks']
    for c, h in enumerate(headers, 1): ws.cell(2, c).value = h
    style_header(ws, 2, len(headers))

    for col in range(5, 5 + len(subjects)):
        ws.cell(row=2, column=col).fill = PatternFill(start_color='1D6A3A', end_color='1D6A3A', fill_type='solid')
        ws.cell(row=2, column=col).font = Font(bold=True, color='FFFFFF', size=11)
        
    calc_start = 5 + len(subjects)
    for col in range(calc_start, calc_start + 5):
        ws.cell(row=2, column=col).fill = PatternFill(start_color='0C4A6E', end_color='0C4A6E', fill_type='solid')
        ws.cell(row=2, column=col).font = Font(bold=True, color='FFFFFF', size=11)

    rows_data = []
    for s in filtered:
        prog = s.get('examination_progress', [])
        exam_entry = next((e for e in prog if (e.get('exam_name') or e.get('exam','')) == exam), None)
        
        scores = {sub: '' for sub in subjects}
        total = ''
        pct = ''
        grade = ''
        rank = 9999
        result = ''
        remarks = ''
        
        if exam_entry:
            for sub_obj in exam_entry.get('subjects', []):
                sname = sub_obj.get('name')
                if sname in scores:
                    scores[sname] = sub_obj.get('obtained', 0)
            total = exam_entry.get('total', 0)
            pct = exam_entry.get('percentage', 0)
            grade = exam_entry.get('grade', '')
            result = 'FAIL' if pct < 35 else 'PASS'
            remarks = 'Needs Improvement' if pct < 35 else 'Passed'
        
        row_dict = {
            'roll': s['roll_no'],
            'name': s['name'],
            'class': s['class'],
            'section': s['section'],
            'scores': scores,
            'total': total,
            'pct': pct,
            'grade': grade,
            'rank': rank,
            'result': result,
            'remarks': remarks
        }
        rows_data.append(row_dict)

    # Dynamic ranking
    scored_students = [r for r in rows_data if r['pct'] != '']
    scored_students.sort(key=lambda x: -x['pct'])
    for idx, r in enumerate(scored_students, start=1):
        r['rank'] = idx
        r['remarks'] = 'Needs Improvement' if r['pct'] < 35 else 'Passed'

    rows_data.sort(key=lambda x: x['rank'])

    grade_colors = {'A+':'D4EDDA','A':'C3E6CB','B+':'D1ECF1','B':'BEE5EB',
                    'C':'FFF3CD','D':'FFE8A1','F':'F8D7DA',''  :'FFFFFF'}

    for r_i, row in enumerate(rows_data, start=3):
        sub_vals = [row['scores'][sub] for sub in subjects]
        rank_label = f"#{row['rank']}" if row['rank'] != 9999 else ''
        pct_label = f"{row['pct']}%" if row['pct'] != '' else ''
        
        vals = [row['roll'], row['name'], row['class'], row['section']] + sub_vals + \
               [row['total'], pct_label, row['grade'], rank_label, row['result'], row['remarks']]
        
        even = r_i%2==0
        for c, val in enumerate(vals, 1):
            cell = ws.cell(r_i, c, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CCCCCC')
            cell.border = Border(left=thin,right=thin,top=thin,bottom=thin)
            
            # Highlight Grade and Result columns
            grade_col_idx = 4 + len(subjects) + 3
            result_col_idx = 4 + len(subjects) + 5
            rank_col_idx = 4 + len(subjects) + 4
            
            if c == grade_col_idx:
                gc = grade_colors.get(row['grade'], 'FFFFFF')
                cell.fill = PatternFill(start_color=gc, end_color=gc, fill_type='solid')
                cell.font = Font(bold=True)
            elif c == result_col_idx and row['result']:
                rc = 'D4EDDA' if row['result'] == 'PASS' else 'F8D7DA'
                tc = '065F46' if row['result'] == 'PASS' else '842029'
                cell.fill = PatternFill(start_color=rc, end_color=rc, fill_type='solid')
                cell.font = Font(bold=True, color=tc)
            elif c == rank_col_idx:
                cell.font = Font(bold=True, color='1E3A5F')
            else:
                bg = 'F8F9FA' if even else 'FFFFFF'
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')

    # Calculate column widths dynamically
    col_widths = [10, 24, 8, 10] + [14] * len(subjects) + [10, 12, 8, 12, 10, 16]
    for c, w in enumerate(col_widths, 1): 
        ws.column_dimensions[get_column_letter(c)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f'SPTAS_Ranks_Report_Class{cls or "All"}{sec or ""}_{exam.replace(" ","_")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =========================================================
# DOWNLOAD MONTHLY ATTENDANCE EXCEL REPORT
# =========================================================
@app.route('/api/excel/attendance-report', methods=['GET'])
def download_attendance_report():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500
    cls = request.args.get('class', '')
    sec = request.args.get('section', '')
    month = request.args.get('month', '')
    
    if not month:
        return jsonify({"error": "Month is required"}), 400
        
    teacher_id = request.headers.get('X-Teacher-Id', '')
    allowed_classes = get_teacher_classes(teacher_id)
    if allowed_classes is not None:
        if cls and sec:
            s_key = f"{cls}-{sec}"
            if s_key not in allowed_classes:
                return jsonify({"error": "Forbidden: This is not your Class"}), 403
        elif cls:
            matched = [ac for ac in allowed_classes if ac.split('-')[0] == cls]
            if not matched:
                return jsonify({"error": "Forbidden: This is not your Class"}), 403
        else:
            return jsonify({"error": "Forbidden: Teacher must specify an assigned Class"}), 403

    students = load_db()
    filtered = sorted([s for s in students if (not cls or str(s['class'])==cls) and
                (not sec or s['section']==sec)], key=lambda x: x['roll_no'])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance Report'
    ws.freeze_panes = 'E3'

    # Title banner
    ws.merge_cells('A1:J1')
    ws['A1'] = f'SPTAS — Monthly Attendance Report | Class {cls or "All"}-{sec or "All"} | Month: {month}'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['Roll No', 'Student Name', 'Class', 'Section', 'Present Days', 'Absent Days', 'Half Days', 'Total Work Days', 'Attendance %', 'Status']
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx, value=h)
    style_header(ws, 2, len(headers))

    # Add rows
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 3
    for s in filtered:
        recs = [r for r in s.get("attendance_records", []) if r["date"].startswith(month)]
        workdays = sum(1 for r in recs if r["status"] != "Holiday")
        present = sum(1 for r in recs if r["status"] == "Present")
        absent = sum(1 for r in recs if r["status"] == "Absent")
        half = sum(1 for r in recs if r["status"] == "Half Day")
        effective_present = present + (half / 2.0)
        percentage = round((effective_present / workdays) * 100, 1) if workdays > 0 else 0.0
        status_str = "Good Standing" if percentage >= 75 else "Shortage"

        vals = [
            s['roll_no'],
            s['name'],
            s['class'],
            s['section'],
            present,
            absent,
            half,
            workdays,
            f"{percentage}%" if workdays > 0 else "0.0%",
            status_str
        ]

        even = row_idx % 2 == 0
        bg_color = 'F8F9FA' if even else 'FFFFFF'
        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        for col_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.fill = row_fill
            if col_idx == 10:
                sc = 'D4EDDA' if percentage >= 75 else 'F8D7DA'
                st = '065F46' if percentage >= 75 else '842029'
                cell.fill = PatternFill(start_color=sc, end_color=sc, fill_type='solid')
                cell.font = Font(bold=True, color=st)

        row_idx += 1

    # Set column widths
    col_widths = [10, 24, 8, 10, 14, 14, 14, 16, 16, 16]
    for c_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f'SPTAS_Attendance_Report_Class{cls or "All"}{sec or ""}_{month}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/teacher/profile/<tid>', methods=['GET'])
def get_teacher_profile(tid):
    for t in load_teachers():
        if t["id"] == tid:
            profile = dict(t)
            classes = profile.get("classes")
            if not classes and profile.get("class") and profile.get("section"):
                profile["classes"] = [f"{profile['class']}-{profile['section']}"]
            elif not classes:
                profile["classes"] = []
            return jsonify(profile)
    return jsonify({"error": "Not found"}), 404


# =========================================================
# DOWNLOAD STUDENT REGISTRATION SAMPLE TEMPLATE
# =========================================================
@app.route('/api/excel/sample-register', methods=['GET'])
def download_sample_register():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'openpyxl not installed'}), 500

    cls = request.args.get('class', '10')
    if not cls or cls.lower() == 'all':
        cls = '10'
    sec = request.args.get('section', 'A')
    if not sec or sec.lower() == 'all':
        sec = 'A'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Student Registration'
    ws.freeze_panes = 'C3'

    # Title Banner
    total_cols = 11
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f'SPTAS — Bulk Student Registration Template'
    title_cell.font = Font(bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Header Row
    headers = [
        'S.No', 'Student Name', 'Roll No', 'Admission No', 'Class', 'Section',
        'Parent Name', 'Parent Contact', 'Parent Alt Contact', 'DOB',
        'Academic Year'
    ]
    
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=2, column=col_idx).value = h
    style_header(ws, 2, len(headers), fill_hex='1E3A5F')

    # Let's add 2 example rows
    examples = [
        (1, 'Rahul Kumar', '101', 'ADM001', cls, sec, 'Srinivasa Rao', '7013029211', '9876543211', '2015-05-15', '2025-26'),
        (2, 'Aditya Vardhan', '102', 'ADM002', cls, sec, 'Kalyan Rao', '9876543220', '', '2015-08-20', '2025-26')
    ]
    
    data_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, row_val in enumerate(examples, start=3):
        for col_idx, val in enumerate(row_val, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            cell.fill = data_fill

    # Set column widths
    col_widths = [6, 22, 10, 15, 8, 10, 20, 15, 18, 12, 14]
    for c_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'SPTAS_Student_Register_Class{cls}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# =========================================================
# UPLOAD & PROCESS STUDENT REGISTRATION EXCEL
# =========================================================
@app.route('/api/excel/upload-register', methods=['POST'])
def upload_register_excel():
    if not OPENPYXL_AVAILABLE:
        return jsonify({'success': False, 'message': 'openpyxl not available'}), 500
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400

    f = request.files['file']
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
    except Exception as e:
        return jsonify({'success': False, 'message': f'Cannot read file: {str(e)}'}), 400

    ws = wb.active

    # Read header row (row 2)
    headers = []
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=2, column=col).value
        headers.append(str(h).strip().lower() if h else '')

    # Try mapping column names dynamically
    col_map = {}
    standard_fields = {
        'student name': ['student name', 'name', 'student'],
        'roll no': ['roll no', 'roll number', 'roll_no', 'roll'],
        'admission no': ['admission no', 'admission number', 'admission_no', 'admission'],
        'class': ['class', 'grade'],
        'section': ['section', 'sec'],
        'parent name': ['parent name', 'father name', 'mother name', 'parent_name', 'parent'],
        'parent contact': ['parent contact', 'phone', 'contact', 'mobile', 'parent_contact'],
        'parent alt contact': ['parent alt contact', 'alt contact', 'secondary contact', 'parent_alt_contact'],
        'dob': ['dob', 'date of birth', 'birth'],
        'academic year': ['academic year', 'year', 'academic_year'],
        'class teacher': ['class teacher', 'teacher', 'class_teacher'],
        'attendance percentage': ['attendance %', 'attendance percentage', 'attendance']
    }

    # Find where each field maps
    for std_field, aliases in standard_fields.items():
        found_col = None
        for alias in aliases:
            if alias in headers:
                found_col = headers.index(alias) + 1
                break
        col_map[std_field] = found_col

    # Check required fields mapping
    required_mappings = ['student name', 'roll no', 'class', 'section', 'parent contact']
    missing_mappings = [m for m in required_mappings if col_map[m] is None]
    if missing_mappings:
        return jsonify({
            'success': False,
            'message': f"Could not find columns for required fields in the Excel headers: {', '.join(missing_mappings)}. Check column headings."
        }), 400

    # Collect subject columns (any column not matching standard or S.No/sno)
    subject_cols = []
    all_std_aliases = [alias for aliases in standard_fields.values() for alias in aliases] + ['s.no', 'sno', 'serial']
    for idx, h in enumerate(headers, 1):
        if h and h not in all_std_aliases:
            # Reconstruct neat name (e.g. capitalize)
            subj_name = ws.cell(row=2, column=idx).value
            subject_cols.append((idx, subj_name))

    import time
    students = load_db()
    timetables = load_timetables()
    
    # Store dynamic validation results
    added = 0
    errors = []
    imported_students_summary = []

    # Helper to retrieve value safely from col_map
    def get_cell_val(row_idx, field_name, default=""):
        col_idx = col_map.get(field_name)
        if col_idx is None:
            return default
        val = ws.cell(row=row_idx, column=col_idx).value
        return val if val is not None else default

    for row in range(3, ws.max_row + 1):
        # Verify if row is empty by checking if name is empty
        name_val = get_cell_val(row, 'student name', None)
        if not name_val:
            continue
        
        name = str(name_val).strip()
        roll_no = str(get_cell_val(row, 'roll no', '')).strip()
        s_class = str(get_cell_val(row, 'class', '10')).strip()
        s_section = str(get_cell_val(row, 'section', 'A')).strip().upper()
        parent_contact = clean_phone(get_cell_val(row, 'parent contact', ''))

        # Basic validations
        if not name or not roll_no or not s_class or not s_section or not parent_contact:
            errors.append(f"Row {row}: Missing required data (Name, Roll, Class, Section, or Parent Contact).")
            continue

        # Check duplication in existing DB
        if any(s['class'] == s_class and s['section'] == s_section and s['roll_no'] == roll_no for s in students):
            errors.append(f"Row {row}: Student '{name}' with Roll {roll_no} already exists in Class {s_class}-{s_section}.")
            continue

        # Parse dob
        dob_val = get_cell_val(row, 'dob', None)
        dob_str = ""
        if dob_val:
            if isinstance(dob_val, datetime):
                dob_str = dob_val.strftime('%Y-%m-%d')
            else:
                dob_str = str(dob_val).strip()

        # Parse other info
        admission_no = str(get_cell_val(row, 'admission no', '')).strip()
        parent_name = str(get_cell_val(row, 'parent name', '')).strip()
        parent_alt_contact = clean_phone(get_cell_val(row, 'parent alt contact', ''))
        academic_year = str(get_cell_val(row, 'academic year', '2025-26')).strip()
        class_teacher = str(get_cell_val(row, 'class teacher', 'Class Teacher')).strip()
        
        att_pct_val = get_cell_val(row, 'attendance percentage', None)
        try: att_pct = float(att_pct_val) if att_pct_val is not None else 90.0
        except: att_pct = 90.0

        # Parse subjects and baseline test marks
        student_subjects = [s_name for idx, s_name in subject_cols]
        if not student_subjects:
            student_subjects = ['Telugu', 'Hindi', 'Mathematics', 'Science', 'Social Studies']

        # Construct examination baseline marks
        exam_subjects = []
        subject_performance = []
        tot_obtained = 0
        tot_max = 0
        
        for col_idx, s_name in subject_cols:
            score_val = ws.cell(row=row, column=col_idx).value
            try:
                score = float(score_val) if score_val is not None else 0.0
            except:
                score = 0.0
            
            exam_subjects.append({
                'name': s_name,
                'obtained': score,
                'max': 100
            })
            subject_performance.append({
                'subject': s_name,
                'score': int(score)
            })
            tot_obtained += score
            tot_max += 100

        # Compile baseline exam record
        baseline_pct = round((tot_obtained / tot_max) * 100, 2) if tot_max > 0 else 0
        baseline_grade = get_grade(baseline_pct)
        
        examination_progress = []
        if subject_cols:
            examination_progress.append({
                'exam_name': 'Baseline Test',
                'exam': 'Baseline Test',
                'year': str(datetime.now().year),
                'subjects': exam_subjects,
                'total': tot_obtained,
                'total_max': tot_max,
                'percentage': baseline_pct,
                'grade': baseline_grade,
                'rank': 1
            })

        # Apply class timetable
        class_key = f"{s_class}-{s_section}"
        class_timetable = timetables.get(class_key, {})

        # Generate unique ID
        import random
        rand_suffix = ''.join(random.choices(string.digits, k=4))
        student_id = f"{int(time.time())}{rand_suffix}"

        new_student = {
            "id": student_id,
            "name": name,
            "roll_no": roll_no,
            "admission_no": admission_no if admission_no else f"ADM{student_id[-5:]}",
            "class": s_class,
            "section": s_section,
            "academic_year": academic_year,
            "dob": dob_str,
            "parent_name": parent_name if parent_name else f"Parent of {name}",
            "parent_contact": parent_contact,
            "parent_alt_contact": parent_alt_contact,
            "parent_password": "",
            "parent_feedback": "",
            "principal_reply": "",
            "attendance_status": "Present",
            "attendance_records": [],
            "fees": {
                "school": 10000,
                "tuition": 30000,
                "books": 5000,
                "dresses": 3000,
                "extra": 0,
                "paid": 0
            },
            "current_status": {
                "standard": f"Class {s_class}",
                "section": f"Section {s_section}",
                "class_teacher": class_teacher,
                "attendance_percentage": att_pct,
                "subjects": student_subjects
            },
            "timetable": class_timetable,
            "examination_progress": examination_progress,
            "subject_performance": subject_performance,
            "progress_comparison": [],
            "performance_trend": "Stable",
            "teacher_term_remarks": [],
            "behavioral_observation": {"discipline": 4, "leadership": 4, "participation": 4, "communication": 4, "teamwork": 4, "confidence": 4},
            "co_curricular_activities": [],
            "awards": [],
            "parent_meetings": []
        }
        # Collect parsed students
        imported_students_summary.append(new_student)
        added += 1

    return jsonify({
        'success': True,
        'preview': True,
        'type': 'register',
        'data': imported_students_summary,
        'errors': errors
    })


# =========================================================
# SAVE/CONFIRM EXCEL IMPORTS (Student registration, Marks, Attendance)
# =========================================================
@app.route('/api/excel/confirm-import', methods=['POST'])
def confirm_import():
    data_payload = request.get_json() or {}
    import_type = data_payload.get('type')
    records = data_payload.get('data', [])
    if not import_type or not records:
        return jsonify({'success': False, 'message': 'Missing data or import type'}), 400

    students = load_db()
    student_map = {s['roll_no']: s for s in students}
    updated = 0
    role, name = get_log_identity()

    if import_type == 'register':
        import time
        import random
        import string
        timetables = load_timetables()
        added_records = []
        for r in records:
            s_class = r.get('class')
            s_sec = r.get('section')
            roll = r.get('roll_no')
            # Check duplication inside DB or inside current import list
            if any(s['class'] == s_class and s['section'] == s_sec and s['roll_no'] == roll for s in students):
                continue
            
            # Re-generate IDs to ensure no collisions
            rand_suffix = ''.join(random.choices(string.digits, k=4))
            student_id = f"{int(time.time())}{rand_suffix}"
            r['id'] = student_id
            
            # Apply class timetable
            class_key = f"{s_class}-{s_sec}"
            r['timetable'] = timetables.get(class_key, {})
            r['teacher_reply'] = ''  # default empty teacher reply
            
            students.append(r)
            updated += 1
            added_records.append({'name': r['name'], 'roll_no': roll, 'class': s_class, 'section': s_sec})
        
        if save_db(students):
            log_activity(role, name, f"Imported and Registered {updated} Students via Excel")
            return jsonify({'success': True, 'updated': updated, 'results': added_records})

    elif import_type == 'marks':
        results_summary = []
        for r in records:
            roll = r.get('roll_no')
            s = student_map.get(roll)
            if not s: continue
            exam_name = r.get('exam')
            max_marks = r.get('max_marks', 100)
            
            exam_entry = {
                'exam_name': exam_name, 'exam': exam_name,
                'year': str(datetime.now().year),
                'subjects': [{'name': sn, 'obtained': sv, 'max': max_marks}
                             for sn, sv in r.get('subjects', {}).items()],
                'total': r.get('total'), 'total_max': r.get('total_max'),
                'percentage': r.get('percentage'),
                'grade': r.get('grade'), 'rank': r.get('rank')
            }
            if 'examination_progress' not in s: s['examination_progress'] = []
            existing_idx = next((i for i, e in enumerate(s['examination_progress'])
                                 if (e.get('exam_name') or e.get('exam','')) == exam_name), None)
            if existing_idx is not None:
                s['examination_progress'][existing_idx] = exam_entry
            else:
                s['examination_progress'].append(exam_entry)

            # Update subject_performance
            s['subject_performance'] = [{'subject': sn, 'score': round((sv / max_marks) * 100)}
                                         for sn, sv in r.get('subjects', {}).items()]
            
            # Update performance trend
            prog = s['examination_progress']
            if len(prog) >= 2:
                pcts = [e.get('percentage', 0) for e in prog[-2:]]
                diff = pcts[-1] - pcts[-2]
                s['performance_trend'] = 'Improving' if diff > 2 else ('Declining' if diff < -2 else 'Stable')

            updated += 1
            results_summary.append({
                'roll_no': roll, 'name': s['name'], 'exam': exam_name,
                'total': r.get('total'), 'percentage': r.get('percentage'),
                'grade': r.get('grade'), 'rank': r.get('rank')
            })

        # Recalculate ranks class-wise and exam-wise
        classes_to_update = set()
        for r in records:
            roll = r.get('roll_no')
            s = student_map.get(roll)
            if s:
                classes_to_update.add((s['class'], s['section'], r.get('exam')))
        
        for cls_name, sec_name, exam_name in classes_to_update:
            class_studs = [st for st in students if st['class'] == cls_name and st['section'] == sec_name]
            exam_pcts = []
            for st in class_studs:
                exam_entry = next((e for e in st.get('examination_progress', [])
                                   if (e.get('exam_name') or e.get('exam','')) == exam_name), None)
                if exam_entry:
                    exam_pcts.append((st, exam_entry.get('percentage', 0)))
                else:
                    exam_pcts.append((st, 0))
            
            exam_pcts.sort(key=lambda x: -x[1])
            for rank_idx, (st, pct) in enumerate(exam_pcts, start=1):
                exam_entry = next((e for e in st.get('examination_progress', [])
                                   if (e.get('exam_name') or e.get('exam','')) == exam_name), None)
                if exam_entry:
                    exam_entry['rank'] = rank_idx

        if save_db(students):
            log_activity(role, name, f"Imported Marks for {updated} Students via Excel (Exam: {records[0].get('exam') if records else 'Unknown'})")
            return jsonify({'success': True, 'updated': updated, 'results': results_summary})

    elif import_type == 'attendance':
        for r in records:
            roll = r.get('roll_no')
            s = student_map.get(roll)
            if not s: continue
            if 'attendance_records' not in s: s['attendance_records'] = []
            for rec in r.get('records', []):
                date_str = rec['date']
                status = rec['status']
                existing = next((item for item in s['attendance_records'] if item['date'] == date_str), None)
                if existing: existing['status'] = status
                else: s['attendance_records'].append({'date': date_str, 'status': status})
            
            recs = s.get('attendance_records', [])
            workdays = sum(1 for item in recs if item.get('status') != 'Holiday')
            present_days = sum(1 for item in recs if item.get('status') == 'Present')
            half_days = sum(1 for item in recs if item.get('status') == 'Half Day')
            effective_present = present_days + (half_days / 2.0)
            
            if workdays > 0:
                s['current_status']['attendance_percentage'] = round((effective_present / workdays) * 100, 1)
            else:
                s['current_status']['attendance_percentage'] = 100.0
            updated += 1
            
        if save_db(students):
            log_activity(role, name, f"Imported Attendance for {updated} Students via Excel")
            return jsonify({'success': True, 'updated': updated})

    return jsonify({'success': False, 'message': 'Import failed'}), 500


# =========================================================
# TEACHER FEEDBACK REPLY
# =========================================================
@app.route('/api/teacher/reply', methods=['POST'])
def teacher_reply():
    data = request.get_json() or {}
    sid = data.get("student_id", "")
    reply = data.get("reply", "").strip()
    students = load_db()
    for s in students:
        if s["id"] == sid:
            s["teacher_reply"] = reply
            save_db(students)
            return jsonify({"success": True})
    return jsonify({"success": False}), 404


# =========================================================
# FEES & SERVICES API
# =========================================================
@app.route('/api/students/fees', methods=['PUT'])
def update_student_fees():
    role, name = get_log_identity()
    if role.lower() != "admin":
        return jsonify({"success": False, "message": "Access denied: Administrator role required."}), 403
        
    data = request.get_json() or {}
    student_id = data.get("student_id")
    if not student_id:
        return jsonify({"success": False, "message": "student_id is required."}), 400
        
    students = load_db()
    student = None
    for s in students:
        if s["id"] == student_id:
            student = s
            break
            
    if not student:
        return jsonify({"success": False, "message": "Student not found."}), 404
        
    fees = student.get("fees", {})
    fees["school"] = int(data.get("school", fees.get("school", 10000)))
    fees["tuition"] = int(data.get("tuition", fees.get("tuition", 30000)))
    fees["books"] = int(data.get("books", fees.get("books", 5000)))
    fees["dresses"] = int(data.get("dresses", fees.get("dresses", 3000)))
    fees["extra"] = int(data.get("extra", fees.get("extra", 0)))
    fees["paid"] = int(data.get("paid", fees.get("paid", 0)))
    student["fees"] = fees
    
    if save_db(students):
        log_activity(role, name, f"Updated fees for student {student['name']} ({student_id})")
        return jsonify({"success": True, "message": "Fees updated successfully."})
    return jsonify({"success": False, "message": "Failed to save to database."}), 500


@app.route('/api/services', methods=['GET'])
def get_services():
    return jsonify(load_json('services_db.json', []))


@app.route('/api/services/add', methods=['POST'])
def add_service():
    role, name = get_log_identity()
    if role.lower() not in ["admin", "principal"]:
        return jsonify({"success": False, "message": "Access denied."}), 403
    data = request.get_json() or {}
    s_name = data.get("name", "").strip()
    s_role = data.get("role", "").strip()
    s_phone = data.get("phone", "").strip()
    if not s_name or not s_role or not s_phone:
        return jsonify({"success": False, "message": "All fields are required."}), 400
        
    import time
    services = load_json('services_db.json', [])
    new_s = {
        "id": str(int(time.time())) + str(random.randint(10, 99)),
        "name": s_name,
        "role": s_role,
        "phone": s_phone
    }
    services.append(new_s)
    if save_json('services_db.json', services):
        log_activity(role, name, f"Added service contact: {s_name} ({s_role})")
        return jsonify({"success": True, "message": "Service contact added."})
    return jsonify({"success": False, "message": "Failed to save."}), 500


@app.route('/api/services/delete/<sid>', methods=['POST', 'DELETE'])
def delete_service(sid):
    role, name = get_log_identity()
    if role.lower() not in ["admin", "principal"]:
        return jsonify({"success": False, "message": "Access denied."}), 403
    services = load_json('services_db.json', [])
    filtered = [s for s in services if s["id"] != sid]
    if len(filtered) == len(services):
        return jsonify({"success": False, "message": "Contact not found."}), 404
    if save_json('services_db.json', filtered):
        log_activity(role, name, f"Deleted service contact ID {sid}")
        return jsonify({"success": True, "message": "Service contact deleted."})
    return jsonify({"success": False, "message": "Failed to save."}), 500





if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)

